''' Library to manage information by the database '''

#Django imports
from django.shortcuts import render,redirect

#Personal imports
from employees.models import Employees, Attendance,Incidences
import xlrd
import datetime
from datetime import date
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from PyPDF3 import PdfFileMerger

from django.http.response import HttpResponse

def merge_payroll(excel_files):
    pass
    merger = PdfFileMerger()
    for i in excel_files:
        pass
        print(i.temporary_file_path())
        the_file = open(str(i.temporary_file_path()), "rb")
        merger.append(the_file)
    try:
        os.mkdir("C:/tempx")
    except:
        pass
    output = open("C:/tempx/archivo.pdf", "wb")
    merger.write(output)
    path="C:/tempx/archivo.pdf"
    os.startfile(path)

def generate_payroll(excel_file):
    pass

    wb = xlrd.open_workbook(excel_file.temporary_file_path())
    wb = load_workbook(filename=excel_file.temporary_file_path(), data_only=True)
    ws_data = wb["INGRESO DE DATOS"]
    ws_payroll_movements = wb["MOVIMIENTOS DE NOMINA"]
    ws_absences = wb["FALTAS"]
    #We need to obtain the dates from the ws_data
    monday=ws_data.cell(4, 130).value
    thuesday=ws_data.cell(4, 131).value
    wednesday=ws_data.cell(4, 132).value
    thursday=ws_data.cell(4, 133).value
    friday=ws_data.cell(4, 134).value
    saturday=ws_data.cell(4, 135).value
    #Create two excel files to work with them
    wb_absences_cr,wb_incidences_cr = create_payroll_excels()
    ws_absences_cr = wb_absences_cr.active
    ws_incidences_cr = wb_incidences_cr.active
    #Counters
    counter_absences_row=1
    counter_incidences_row = 2
    mrow= ws_absences.max_row
    for x in range(2,mrow-1):
        pass
        if(ws_absences.cell(row=x, column=1).value!=0):
            worker_number = ws_absences.cell(row=x, column=1).value
            #print(ws_absences.cell(row=x, column=4).value)
            porcentaje_l= int(round(float(ws_absences.cell(row=x, column=5).value)*100))
            porcentaje_ma=int(round(float(ws_absences.cell(row=x, column=6).value)*100))
            porcentaje_mi=int(round(float(ws_absences.cell(row=x, column=7).value)*100))
            porcentaje_ju=int(round(float(ws_absences.cell(row=x, column=8).value)*100))
            porcentaje_vi=int(round(float(ws_absences.cell(row=x, column=9).value)*100))
            porcentaje_sab=int(round(float(ws_absences.cell(row=x, column=10).value)*100))

            if porcentaje_l>0:
                pass
                counter_absences_row=counter_absences_row+1
                ws_absences_cr.cell(row=counter_absences_row, column=1).value = worker_number
                ws_absences_cr.cell(row=counter_absences_row, column=2).value = 2
                ws_absences_cr.cell(row=counter_absences_row, column=4).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=5).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=6).value = monday
                ws_absences_cr.cell(row=counter_absences_row, column=7).value = monday

                ws_absences_cr.cell(row=counter_absences_row, column=3).value = porcentaje_l
            if porcentaje_ma>0:
                pass
                counter_absences_row = counter_absences_row + 1
                ws_absences_cr.cell(row=counter_absences_row, column=3).value = porcentaje_ma
                ws_absences_cr.cell(row=counter_absences_row, column=1).value = worker_number
                ws_absences_cr.cell(row=counter_absences_row, column=2).value = 2
                ws_absences_cr.cell(row=counter_absences_row, column=4).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=5).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=6).value = thuesday
                ws_absences_cr.cell(row=counter_absences_row, column=7).value = thuesday

                ws_absences_cr.cell(row=counter_absences_row, column=3).value = porcentaje_ma
            if porcentaje_mi>0:
                pass
                counter_absences_row = counter_absences_row + 1
                ws_absences_cr.cell(row=counter_absences_row, column=3).value = porcentaje_mi
                ws_absences_cr.cell(row=counter_absences_row, column=1).value = worker_number
                ws_absences_cr.cell(row=counter_absences_row, column=2).value = 2
                ws_absences_cr.cell(row=counter_absences_row, column=4).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=5).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=6).value = wednesday
                ws_absences_cr.cell(row=counter_absences_row, column=7).value = wednesday

                ws_absences_cr.cell(row=counter_absences_row, column=3).value = porcentaje_mi
            if porcentaje_ju>0:
                pass
                counter_absences_row = counter_absences_row + 1
                ws_absences_cr.cell(row=counter_absences_row, column=3).value = porcentaje_ju
                ws_absences_cr.cell(row=counter_absences_row, column=1).value = worker_number
                ws_absences_cr.cell(row=counter_absences_row, column=2).value = 2
                ws_absences_cr.cell(row=counter_absences_row, column=4).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=5).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=6).value = thursday
                ws_absences_cr.cell(row=counter_absences_row, column=7).value = thursday
            if porcentaje_vi>0:
                pass
                counter_absences_row = counter_absences_row + 1
                ws_absences_cr.cell(row=counter_absences_row, column=3).value = porcentaje_vi
                ws_absences_cr.cell(row=counter_absences_row, column=1).value = worker_number
                ws_absences_cr.cell(row=counter_absences_row, column=2).value = 2
                ws_absences_cr.cell(row=counter_absences_row, column=4).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=5).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=6).value = friday
                ws_absences_cr.cell(row=counter_absences_row, column=7).value = friday
            if porcentaje_sab>0:
                pass
                counter_absences_row = counter_absences_row + 1
                ws_absences_cr.cell(row=counter_absences_row, column=3).value = porcentaje_sab
                ws_absences_cr.cell(row=counter_absences_row, column=1).value = worker_number
                ws_absences_cr.cell(row=counter_absences_row, column=2).value = 2
                ws_absences_cr.cell(row=counter_absences_row, column=4).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=5).value = 0
                ws_absences_cr.cell(row=counter_absences_row, column=6).value = saturday
                ws_absences_cr.cell(row=counter_absences_row, column=7).value = saturday

    nrow = ws_payroll_movements.max_row
    print("TOTAL ROWS: "+str(nrow))
    cont_mov_created=1
    for x in range(4, nrow+1):
        pass
        worker_number = str(ws_payroll_movements.cell(row=x, column=1).value)
        if worker_number !="0":
            print("Clave:    "+str(worker_number))
            print("Valores 5:"+str(ws_payroll_movements.cell(row=x, column=5).value)+"||")
            cont_mov_created=cont_mov_created+1
            print(ws_payroll_movements.cell(row=x, column=5).value)
            print("///////////////////")
            
            print_payroll_movements(cont_mov_created, ws_incidences_cr, worker_number, "P","1",saturday, ws_payroll_movements.cell(row=x, column=5).value)
            print("lol")
            # SEPTIMO DÍA P109
            # print("Valores 6:" + str(ws_payroll_movements.cell(row=x, column=6).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=6).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "109", saturday,
                                                ws_payroll_movements.cell(row=x, column=6).value)

            #MONTO HORAS DOBLES P003
            # print("Valores 7:" + str(ws_payroll_movements.cell(row=x, column=7).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=7).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "3", saturday,
                                                ws_payroll_movements.cell(row=x, column=7).value)

            #MONTO HORAS TRIPLES P005
            # print("Valores 8:" + str(ws_payroll_movements.cell(row=x, column=8).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=8).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "5", saturday,
                                                ws_payroll_movements.cell(row=x, column=8).value)

            #VACACIONES P009
            # print("Valores 9:" + str(ws_payroll_movements.cell(row=x, column=9).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=9).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "9", saturday,
                                                ws_payroll_movements.cell(row=x, column=9).value)

            #PRIMA VACACIONAL P010
            # print("Valores 10:" + str(ws_payroll_movements.cell(row=x, column=10).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=10).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "10", saturday,
                                                ws_payroll_movements.cell(row=x, column=10).value)

            #DESCANSO LABORADO P118
            # print("Valores 11:" + str(ws_payroll_movements.cell(row=x, column=11).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=11).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "118", saturday,
                                                ws_payroll_movements.cell(row=x, column=11).value)

            #PRIMA DOMINICAL P019
            # print("Valores 12:" + str(ws_payroll_movements.cell(row=x, column=12).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=12).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "19", saturday,
                                                ws_payroll_movements.cell(row=x, column=12).value)

            #BONO POR CAPACITACION P113
            # print("Valores 13:" + str(ws_payroll_movements.cell(row=x, column=13).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=13).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "113", saturday,
                                                ws_payroll_movements.cell(row=x, column=13).value)

            # BONO ÁREA ESPECIAL P121
            # print("Valores 14:" + str(ws_payroll_movements.cell(row=x, column=14).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=14).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "121", saturday,
                                                ws_payroll_movements.cell(row=x, column=14).value)
            #BONO LIDER DE GRUPO P129
            # print("Valores 15: " + str(ws_payroll_movements.cell(row=x, column=15).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=15).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "129", saturday,
                                                ws_payroll_movements.cell(row=x, column=15).value)

            #BONO POR RENDIMIENTO P132
            # print("Valores: 16 " + str(ws_payroll_movements.cell(row=x, column=16).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=16).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "132", saturday,
                                                ws_payroll_movements.cell(row=x, column=16).value)

            #PREMIOS DE EFICIENCIA P126
            # print("Valores 17:" + str(ws_payroll_movements.cell(row=x, column=17).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=17).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "126", saturday,
                                                ws_payroll_movements.cell(row=x, column=17).value)

            #PREMIOS DE PUNTUALIDAD P111
            # print("Valores: 18" + str(ws_payroll_movements.cell(row=x, column=18).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=18).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "111", saturday,
                                                ws_payroll_movements.cell(row=x, column=18).value)

            #BONO KPI P135
            # print("Valores 19:" + str(ws_payroll_movements.cell(row=x, column=19).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=19).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "135", saturday,
                                                ws_payroll_movements.cell(row=x, column=19).value)

            #BONO POR ESPECIALIDAD P 133
            # print("Valores 20:" + str(ws_payroll_movements.cell(row=x, column=20).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=20).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "133", saturday,
                                                ws_payroll_movements.cell(row=x, column=20).value)
            #APOYO A ENCARGADO P134
            # print("Valores: 21" + str(ws_payroll_movements.cell(row=x, column=21).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=21).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "134", saturday,
                                                ws_payroll_movements.cell(row=x, column=21).value)

            #OTROS BONOS P114
            # print("Valores 22:" + str(ws_payroll_movements.cell(row=x, column=22).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=22).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "114", saturday,
                                                ws_payroll_movements.cell(row=x, column=22).value)
            print(str(worker_number)+" - "+str(ws_payroll_movements.cell(row=x, column=22).value))

            #OTROS BONOS 2 P130
            # print("Valores 23:" + str(ws_payroll_movements.cell(row=x, column=23).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=23).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "130", saturday,
                                                ws_payroll_movements.cell(row=x, column=23).value)

            #INCENTIVO DE PRODUCTIVIDAD P131
            # print("Valores 24:" + str(ws_payroll_movements.cell(row=x, column=24).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=24).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "131", saturday,
                                                ws_payroll_movements.cell(row=x, column=24).value)

            #RETROACTIVO P127
            # print("Valores 25:" + str(ws_payroll_movements.cell(row=x, column=25).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=25).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                                worker_number, "P", "127", saturday,
                                                ws_payroll_movements.cell(row=x, column=25).value)

            # RETROACTIVO P127
            # print("Valores 25:" + str(ws_payroll_movements.cell(row=x, column=25).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=25).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements(cont_mov_created, ws_incidences_cr,
                                            worker_number, "D", "113", saturday,
                                            ws_payroll_movements.cell(row=x, column=26).value)


            #REPOSICION GAFETE D114
            # print("Valores 27:" + str(ws_payroll_movements.cell(row=x, column=27).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=27).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements( cont_mov_created, ws_incidences_cr,
                                            worker_number, "D", "114", saturday,
                                            ws_payroll_movements.cell(row=x, column=27).value)
            # AGUINALDO P0004
            # print("Valores 27:" + str(ws_payroll_movements.cell(row=x, column=27).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=27).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements(cont_mov_created, ws_incidences_cr,
                                            worker_number, "P", "4", saturday,
                                            ws_payroll_movements.cell(row=x, column=28).value)
            #BONO DE DESPENSA P014
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements(cont_mov_created, ws_incidences_cr,
                                            worker_number, "P", "14", saturday,
                                            ws_payroll_movements.cell(row=x, column=29).value)
            # BONO DE DESEMPENIO P136
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements(cont_mov_created, ws_incidences_cr,
                                            worker_number, "P", "136", saturday,
                                            ws_payroll_movements.cell(row=x, column=30).value)
            # BONO DE DESEMPENIO P137
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements(cont_mov_created, ws_incidences_cr,
                                            worker_number, "P", "137", saturday,
                                            ws_payroll_movements.cell(row=x, column=31).value)

            # RETROACTIVO P127
            # print("Valores 25:" + str(ws_payroll_movements.cell(row=x, column=25).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=25).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements(cont_mov_created, ws_incidences_cr,
                                            worker_number, "P", "138", saturday,
                                            ws_payroll_movements.cell(row=x, column=32).value)

            # PERDIDA DE HERRAMIENTA D113
            # print("Valores 26:" + str(ws_payroll_movements.cell(row=x, column=26).value) + "||")
            # if str(ws_payroll_movements.cell(row=x, column=26).value) != "0":
            pass
            cont_mov_created = cont_mov_created + 1
            print_payroll_movements(cont_mov_created, ws_incidences_cr,
                                            worker_number, "D", "118", saturday,
                                            ws_payroll_movements.cell(row=x, column=33).value)


    ruta=os.getcwd()
    try:
        os.mkdir("C:/tempx")
    except:
        pass
    wb_incidences_cr.save("C:/tempx/"+'out2.xls')
    os.system('start excel.exe '+'C:/tempx/out2.xls')
    wb_absences_cr.save("C:/tempx/"+'out.xls')
    os.system('start excel.exe '+'C:/tempx/out.xls')


    #Valores clave

    #este nadamas es para ver los valores del excel, no hace nada realmente
    """ for columna in range (1,136):
        pass
        if columna==1:
            worker_number =  sheet.cell(row=fila, column=columna).value

        valor = sheet.cell(row=fila, column=columna).value """
    

def print_payroll_movements(fila,sheet,clave_emp,per_ded,codigo,fecha,monto):
    pass
    sheet.cell(row=fila, column=1).value = clave_emp
    sheet.cell(row=fila, column=2).value = per_ded
    sheet.cell(row=fila, column=3).value = codigo
    sheet.cell(row=fila, column=4).value = 0
    sheet.cell(row=fila, column=5).value = "N"
    sheet.cell(row=fila, column=6).value = 2
    sheet.cell(row=fila, column=7).value = fecha
    sheet.cell(row=fila, column=8).value = fecha
    sheet.cell(row=fila, column=9).value = monto


def download_week_atendance(date_1,end_date,wb,ws):
        # Headers of the documents
        ws['A1'] = "DPTO"
        ws['B1'] = "NO"
        ws['C1'] = "NOMBRE"
        ws['D1'] = "EMPRESA"
        ws['E1'] = "EVENTO"
        ws['F1'] = "LUNES"
        ws['G1'] = "NOTA"
        ws['H1'] = "MARTES"
        ws['I1'] = "NOTA"
        ws['J1'] = "MIERCOLES"
        ws['K1'] = "NOTA"
        ws['L1'] = "JUEVES"
        ws['M1'] = "NOTA"
        ws['N1'] = "VIERNES"
        ws['O1'] = "NOTA"
        ws['P1'] = "SABADO"
        ws['Q1'] = "NOTA"
        ws_row = 4
        ws_column = 1
        old_number="9999999"
        actual_number=""
        event=""
        emp_name=""
        emp_group=""
        emp_number=""
        att_date=""
        att_time=""
        #list that save all numbers of employees who has at least 1 day checking on the week
        first_list = []
        second_list=[]
        att_list = Attendance.objects.filter(date__gte=date_1, date__lte=end_date).order_by('employee','date','time')
        for att_elem in att_list:
            att_date=att_elem.date
            att_time=att_elem.time
            event = att_elem.event
            try:
                info = Employees.objects.get(number=att_elem.number)
                actual_number=info.number
                emp_name = info.name
                emp_group= info.group
                emp_number = info.number
            except Exception as e:
                print("No se encontro el usuario, error code: "+ str(e))
                actual_number=""
                emp_name = att_elem.employee
                emp_group= "No encontrado"
                emp_number = att_elem.number
                continue
            
            if old_number not in first_list:
                first_list.append(old_number)

            if old_number!=actual_number:
                ws_row = ws_row+8
                old_number = actual_number

                for x in range(0, 8):
                    ws.cell(ws_row+x, ws_column+1).value = emp_number
                for x in range(0, 5):
                    ws.cell(ws_row+x, ws_column+2).value = emp_name
                for x in range(0, 8):
                    ws.cell(ws_row+x, ws_column+3).value = "NOR" 
            ws.cell(ws_row, ws_column+4).value = "E"
            ws.cell(ws_row+1, ws_column+4).value = "SD"
            ws.cell(ws_row+2, ws_column+4).value = "ED"
            ws.cell(ws_row+3, ws_column+4).value = "S"
            ws.cell(ws_row+4, ws_column+4).value = "E T.E."
            ws.cell(ws_row+5, ws_column+4).value = "S T.E."
            day = datetime.datetime.strptime(
                    str(att_date), '%Y-%m-%d').strftime('%A')
            if day == "Monday":
                if event == "Entrada":
                    ws.cell(ws_row, ws_column+5).value = att_time
                if event == "Salida a descanso":
                    ws.cell(ws_row+1, ws_column+5).value = att_time
                if event == "Regreso descanso":
                    ws.cell(ws_row+2, ws_column+5).value = att_time
                if event == "Salida":
                    ws.cell(ws_row+3, ws_column+5).value = att_time
                if event == "Salida T.E.":
                    ws.cell(ws_row+4, ws_column+5).value = att_time
                if event == "Entrada T.E.":
                    ws.cell(ws_row+5, ws_column+5).value = att_time
            if day == "Tuesday":
                if event == "Entrada":
                    ws.cell(ws_row, ws_column+7).value = att_time
                if event == "Salida a descanso":
                    ws.cell(ws_row+1, ws_column+7).value = att_time
                if event == "Regreso descanso":
                    ws.cell(ws_row+2, ws_column+7).value = att_time
                if event == "Salida":
                    ws.cell(ws_row+3, ws_column+7).value = att_time
                if event == "Salida T.E.":
                    ws.cell(ws_row+4, ws_column+7).value = att_time
                if event == "Entrada T.E.":
                    ws.cell(ws_row+5, ws_column+7).value = att_time
            if day == "Wednesday":
                
                if event == "Entrada":
                    ws.cell(ws_row, ws_column+9).value = att_time
                if event == "Salida a descanso":
                    ws.cell(ws_row+1, ws_column+9).value = att_time
                if event == "Regreso descanso":
                    ws.cell(ws_row+2, ws_column+9).value = att_time
                if event == "Salida":
                    ws.cell(ws_row+3, ws_column+9).value = att_time
                if event == "Salida T.E.":
                    ws.cell(ws_row+4, ws_column+9).value = att_time
                if event == "Entrada T.E.":
                    ws.cell(ws_row+5, ws_column+9).value = att_time
            if day == "Thursday":
                if event == "Entrada":
                    ws.cell(ws_row, ws_column+11).value = att_time
                if event == "Salida a descanso":
                    ws.cell(ws_row+1, ws_column+11).value = att_time
                if event == "Regreso descanso":
                    ws.cell(ws_row+2, ws_column+11).value = att_time
                if event == "Salida":
                    ws.cell(ws_row+3, ws_column+11).value = att_time
                if event == "Salida T.E.":
                    ws.cell(ws_row+4, ws_column+11).value = att_time
                if event == "Entrada T.E.":
                    ws.cell(ws_row+5, ws_column+11).value = att_time
            if day == "Friday":
                if event == "Entrada":
                    ws.cell(ws_row, ws_column+13).value = att_time
                if event == "Salida a descanso":
                    ws.cell(ws_row+1, ws_column+13).value = att_time
                if event == "Regreso descanso":
                    ws.cell(ws_row+2, ws_column+13).value = att_time
                if event == "Salida":
                    ws.cell(ws_row+3, ws_column+13).value = att_time
                if event == "Salida T.E.":
                    ws.cell(ws_row+4, ws_column+13).value = att_time
                if event == "Entrada T.E.":
                    ws.cell(ws_row+5, ws_column+13).value = att_time
            if day == "Saturday":
                if event == "Entrada":
                    ws.cell(ws_row, ws_column+15).value = att_time
                if event == "Salida a descanso":
                    ws.cell(ws_row+1, ws_column+15).value = att_time
                if event == "Regreso descanso":
                    ws.cell(ws_row+2, ws_column+15).value = att_time
                if event == "Salida":
                    ws.cell(ws_row+3, ws_column+15).value = att_time
                if event == "Salida T.E.":
                    ws.cell(ws_row+4, ws_column+15).value = att_time
                if event == "Entrada T.E.":
                    ws.cell(ws_row+5, ws_column+15).value = att_time
        emps = Employees.objects.filter(status="Activo")
        for emp in emps:
            second_list.append(emp.number)
        last_list = set(second_list) - set(first_list)
        for emp_i in last_list:
            if emp_i != 'C00':
                info = Employees.objects.get(number=emp_i)
                ws_row = ws_row+8
                ws.cell(ws_row, ws_column+4).value = "E"
                ws.cell(ws_row+1, ws_column+4).value = "SD"
                ws.cell(ws_row+2, ws_column+4).value = "ED"
                ws.cell(ws_row+3, ws_column+4).value = "S"
                ws.cell(ws_row+4, ws_column+4).value = "E T.E."
                ws.cell(ws_row+5, ws_column+4).value = "S T.E."
                for x in range(0, 8):
                        ws.cell(ws_row+x, ws_column+1).value = info.number
                for x in range(0, 5):
                    ws.cell(ws_row+x, ws_column+2).value = info.name
                for x in range(0, 8):
                    ws.cell(ws_row+x, ws_column+3).value = "NOR"
            
def upload_assistances(excel_file):
    '''Function that will get the assistances from excel,
     and save them on the database'''
    pass
    wb = xlrd.open_workbook(excel_file.temporary_file_path())
    # Getting a particular sheet by name out of many sheets
    ws = wb.sheet_by_index(0)
    offset=1
    for i, row in enumerate(range(ws.nrows)):
        if i < offset:  # (Optionally) skip headers
            continue
        ne = ws.cell(i, 0).value
        a1_as_datetime = datetime.datetime(*xlrd.xldate_as_tuple(ne, wb.datemode))
        a1_as_datetime= str(a1_as_datetime)
        obj = Attendance()
        
        obj.date = a1_as_datetime[0:10]
        obj.time = a1_as_datetime[11:16]

        num = ws.cell(i,1 ).value
        if num.startswith("0"):
            num=num[1:]
        if num.startswith("0"):
            num=num[1:]
            
        if num.startswith("0"):
            num=num[1:]
            
        if num.startswith("0"):
            num=num[1:]
            
        if num.startswith("0"):
            num=num[1:]
            
        if num.startswith("0"):
            num=num[1:]
        obj.number = num
        obj.event = ws.cell(i, 2).value
        obj.device = ws.cell(i, 3).value
        obj.employee = ws.cell(i, 4).value
        obj.save()


def create_employee_excel(wb,ws):
    ''' Function that will get all data from all employeees
    on the databe and create an excel file to import and export data'''
    employees = Employees.objects.raw('SELECT * FROM employees_employees')
    row = 1
    
    for employee in employees.iterator():
        row = row + 1
        insert_on_excel(wb,ws,employee,row)
        
        
def evaluate_str(atributte):
    pass
    not_wanted = ['DoesNotExist',  'MultipleObjectsReturned', '__class__', 
    'clean', 'clean_fields', 'date_error_message', 'delete', 'from_db', 'full_clean',
    'get_deferred_fields', 'get_next_by_created','get_next_by_modified', 'get_previous_by_created',
    'get_previous_by_modified', 'prepare_database_save', 'refresh_from_db',
    'save', 'save_base', 'serializable_value', 'unique_error_message', 'validate_unique']
    for elem in not_wanted:
        if atributte.find(elem)!=-1:
            return False
    return True


pos_name=0
pos_date_monday=0
pos_date_thuesday=0
pos_date_wednesday=0
pos_date_thursday=0
pos_date_friday=0
pos_date_saturday=0
pos_date_sunday=0



def upload_incidences(excel_file):
    wb = xlrd.open_workbook(excel_file.temporary_file_path())
    ws = wb.sheet_by_name("INGRESO DE DATOS")

    
    
    sunday= process_excel_date(ws.cell(3,15).value,wb)
    monday= process_excel_date(ws.cell(3,17).value,wb)
    thuesday= process_excel_date(ws.cell(3,23).value,wb)
    wednesday= process_excel_date(ws.cell(3,29).value,wb)
    thursday = process_excel_date(ws.cell(3,35).value,wb)
    friday= process_excel_date(ws.cell(3,41).value,wb)
    saturday= process_excel_date(ws.cell(3,47).value,wb)

    for j,a in enumerate(range(ws.nrows)):
        #Las asistencias se guardan con un 0 en horas, lo que significa que trabajo 8 horas y 0 horas extra
        #... si tiene horas extra vendra la asistencia con un numero
        #... los minutos que se quitan de esas 8 horas salen del resultado de guardar PS,R,F
        if j>=5:
                employee_number = ws.cell(j,1).value
                employee_number =str(employee_number)
                employee_number=employee_number[:-2]
                employee_name = ws.cell(j,2).value
                #Events passed on sunday
                event_sunday= (ws.cell(j,15).value)
                extras_sunday= str((ws.cell(j,16).value))
                extras_sunday = extras_sunday[:-2]
                #Events passed on monday
                event_monday= str(ws.cell(j,17).value)
                extras_monday=str(ws.cell(j,18).value)
                r_monday= str(ws.cell(j,19).value)
                p_monday= str(ws.cell(j,20).value)
                f_monday= str(ws.cell(j,21).value)
                v_monday= str(ws.cell(j,22).value)
                
                #Events passed on thuesday
                event_thuesday= (ws.cell(j,23).value)
                extras_thuesday= (ws.cell(j,24).value)
                r_thuesday= (ws.cell(j,25).value)
                p_thuesday= (ws.cell(j,26).value)
                f_thuesday= (ws.cell(j,27).value)
                v_thuesday= (ws.cell(j,28).value)
                #Events passed on wednesday
                event_wednesday= (ws.cell(j,29).value)
                extras_wednesday= (ws.cell(j,30).value)
                r_wednesday = (ws.cell(j,31).value)
                p_wednesday= (ws.cell(j,32).value)
                f_wednesday= (ws.cell(j,33).value)
                v_wednesday= (ws.cell(j,34).value)
                #Events passed on thurdsday
                event_thursday= (ws.cell(j,35).value)
                extras_thursday= (ws.cell(j,36).value)
                r_thursday = (ws.cell(j,37).value)
                p_thursday= (ws.cell(j,38).value)
                f_thursday= (ws.cell(j,39).value)
                v_thursday= (ws.cell(j,40).value)
                #Events passed on friday
                event_friday= (ws.cell(j,41).value)
                extras_friday= (ws.cell(j,42).value)
                r_friday = (ws.cell(j,43).value)
                p_friday= (ws.cell(j,44).value)
                f_friday= (ws.cell(j,45).value)
                v_friday= (ws.cell(j,46).value)
                #Events passed on saturday
                event_saturday= (ws.cell(j,47).value)
                extras_saturday= (ws.cell(j,48).value)
                r_saturday = (ws.cell(j,49).value)
                p_saturday= (ws.cell(j,50).value)
                f_saturday = (ws.cell(j,51).value)
                v_saturday = (ws.cell(j,52).value)
                #All events from sunday
                if(event_sunday=="A"):
                    pass
                    if(extras_sunday!=""):
                        upload_incideces(employee_number,employee_name,"A",sunday,extras_sunday,0)

                elif(event_sunday=="DL"):
                    pass
                    upload_incideces(employee_number,employee_name,"DL",sunday,extras_sunday,0)
                #MONDAY
                if(event_monday=="A"):
                    pass
                    if(extras_monday!=""):
                        upload_incideces(employee_number,employee_name,"A",monday,extras_monday,0)
                    else:
                        upload_incideces(employee_number,employee_name,"A",monday,0,0)
                    if(r_monday!=""):
                        upload_incideces(employee_number,employee_name,"R",monday,0,r_monday)
                    if(p_monday!=""):
                        upload_incideces(employee_number,employee_name,"P",monday,0,p_monday)
                    if(f_monday!=""):
                        upload_incideces(employee_number,employee_name,"F",monday,0,f_monday)
                elif(event_monday=="V"):
                    pass
                    if(v_monday!=""):
                        upload_incideces(employee_number,employee_name,"V",monday,v_monday,0)
                elif(event_monday=="DL"):
                    pass
                    upload_incideces(employee_number,employee_name,"DL",monday,extras_monday,0)
                elif(event_monday=="F" or event_monday=="f"):
                    pass
                    upload_incideces(employee_number,employee_name,"F",monday,8,0)
                else:
                    upload_incideces(employee_number,employee_name,event_monday,monday,8,0)
                #THUESDAY
                if(event_thuesday=="A"):
                    pass
                    if(extras_thuesday!=""):
                        upload_incideces(employee_number,employee_name,"A",thuesday,extras_thuesday,0)
                    else:
                        upload_incideces(employee_number,employee_name,"A",thuesday,0,0)
                    if(r_thuesday!=""):
                        upload_incideces(employee_number,employee_name,"R",thuesday,0,r_thuesday)
                    if(p_thuesday!=""):
                        upload_incideces(employee_number,employee_name,"P",thuesday,0,p_thuesday)
                    if(f_thuesday!=""):
                        upload_incideces(employee_number,employee_name,"FP",thuesday,0,f_thuesday)
                elif(event_thuesday=="V"):
                    pass
                    if(v_thuesday!=""):
                        upload_incideces(employee_number,employee_name,"V",thuesday,v_thuesday,0)
                elif(event_thuesday=="DL"):
                    pass
                    upload_incideces(employee_number,employee_name,"DL",thuesday,extras_thuesday,0)
                elif(event_thuesday=="F" or extras_thuesday=="f"):
                    pass
                    upload_incideces(employee_number,employee_name,"F",thuesday,8,0)
                else:
                    upload_incideces(employee_number,employee_name,event_thuesday,thuesday,8,0)
                #WEDNESDAY
                if(event_wednesday=="A"):
                    pass
                    if(extras_wednesday!=""):
                        upload_incideces(employee_number,employee_name,"A",wednesday,extras_wednesday,0)
                    else:
                        upload_incideces(employee_number,employee_name,"A",wednesday,0,0)
                    if(r_wednesday!=""):
                        upload_incideces(employee_number,employee_name,"R",wednesday,0,r_wednesday)
                    if(p_wednesday!=""):
                        upload_incideces(employee_number,employee_name,"P",wednesday,0,p_wednesday)
                    if(f_wednesday!=""):
                        upload_incideces(employee_number,employee_name,"FP",wednesday,0,f_wednesday)
                    if(v_wednesday!=""):
                        upload_incideces(employee_number,employee_name,"A",wednesday,0,v_wednesday)
                elif(event_wednesday=="V"):
                    pass
                    if(v_wednesday!=""):
                        upload_incideces(employee_number,employee_name,"V",wednesday,v_wednesday,0)
                elif(event_wednesday=="DL"):
                    pass
                    upload_incideces(employee_number,employee_name,"DL",wednesday,extras_wednesday,0)
                elif(event_wednesday=="F" or event_wednesday=="f"):
                    pass
                    upload_incideces(employee_number,employee_name,"F",wednesday,8,0)
                else:
                    upload_incideces(employee_number,employee_name,event_wednesday,wednesday,8,0)
                #THURSDAY
                if(event_thursday=="A"):
                    pass
                    if(extras_thursday!=""):
                        upload_incideces(employee_number,employee_name,"A",thursday,extras_thursday,0)
                    else:
                        upload_incideces(employee_number,employee_name,"A",thursday,0,0)
                    if(r_thursday!=""):
                        upload_incideces(employee_number,employee_name,"R",thursday,0,r_thursday)
                    if(p_thursday!=""):
                        upload_incideces(employee_number,employee_name,"P",thursday,0,p_thursday)
                    if(f_thursday!=""):
                        upload_incideces(employee_number,employee_name,"FP",thursday,0,f_thursday)
                    if(v_thursday!=""):
                        upload_incideces(employee_number,employee_name,"A",thursday,0,v_thursday)
                elif(event_thursday=="V"):
                    pass
                    if(v_thursday!=""):
                        upload_incideces(employee_number,employee_name,"V",thursday,v_thursday,0)
                elif(event_thursday=="DL"):
                    pass
                    upload_incideces(employee_number,employee_name,"DL",thursday,extras_thursday,0)
                elif(event_thursday=="F" or event_thursday=="f"):
                    pass
                    upload_incideces(employee_number,employee_name,"F",thursday,8,0)
                else:
                    upload_incideces(employee_number,employee_name,event_thursday,thursday,8,0)
                #FRIDAY
                if(event_friday=="A"):
                    pass
                    if(extras_friday!=""):
                        upload_incideces(employee_number,employee_name,"A",friday,extras_friday,0)
                    else:
                        upload_incideces(employee_number,employee_name,"A",friday,0,0)
                    if(r_friday!=""):
                        upload_incideces(employee_number,employee_name,"R",friday,0,r_friday)
                    if(p_friday!=""):
                        upload_incideces(employee_number,employee_name,"P",friday,0,p_friday)
                    if(f_friday!=""):
                        upload_incideces(employee_number,employee_name,"FP",friday,0,f_friday)
                    if(v_friday!=""):
                        upload_incideces(employee_number,employee_name,"A",friday,0,v_friday)
                elif(event_friday=="V"):
                    pass
                    if(extras_friday!=""):
                        upload_incideces(employee_number,employee_name,"V",friday,extras_friday,0)
                    if(r_friday!=""):
                        upload_incideces(employee_number,employee_name,"V",friday,0,r_friday)
                    if(p_friday!=""):
                        upload_incideces(employee_number,employee_name,"V",friday,0,p_friday)
                    if(f_friday!=""):
                        upload_incideces(employee_number,employee_name,"V",friday,0,f_friday)
                    if(v_friday!=""):
                        upload_incideces(employee_number,employee_name,"V",friday,v_friday,0)
                elif(event_friday=="DL"):
                    pass
                    upload_incideces(employee_number,employee_name,"DL",friday,extras_friday,0)
                elif(event_friday=="F" or event_friday=="f"):
                    pass
                    upload_incideces(employee_number,employee_name,"F",friday,8,0)
                else:
                    upload_incideces(employee_number,employee_name,event_friday,friday,8,0)
                #SATURDAY
                if(event_saturday=="A"):
                    pass
                    if(extras_saturday!=""):
                        upload_incideces(employee_number,employee_name,"A",saturday,extras_saturday,0)
                    else:
                        upload_incideces(employee_number,employee_name,"A",saturday,0,0)
                    if(r_saturday!=""):
                        upload_incideces(employee_number,employee_name,"R",saturday,0,r_saturday)
                    if(p_saturday!=""):
                        upload_incideces(employee_number,employee_name,"P",saturday,0,p_saturday)
                    if(f_saturday!=""):
                        upload_incideces(employee_number,employee_name,"FP",saturday,0,f_saturday)
                    if(v_saturday!=""):
                        upload_incideces(employee_number,employee_name,"A",saturday,0,v_saturday)
                elif(event_saturday=="V"):
                    pass
                    if(extras_saturday!=""):
                        upload_incideces(employee_number,employee_name,"V",saturday,extras_saturday,0)
                    if(r_saturday!=""):
                        upload_incideces(employee_number,employee_name,"V",saturday,0,r_saturday)
                    if(p_saturday!=""):
                        upload_incideces(employee_number,employee_name,"V",saturday,0,p_saturday)
                    if(f_saturday!=""):
                        upload_incideces(employee_number,employee_name,"V",saturday,0,f_saturday)
                    if(v_saturday!=""):
                        upload_incideces(employee_number,employee_name,"V",saturday,v_saturday,0)
                elif(event_saturday=="DL"):
                    pass
                    upload_incideces(employee_number,employee_name,"DL",saturday,extras_saturday,0)
                elif(event_saturday=="F" or event_saturday=="f"):
                    pass
                    upload_incideces(employee_number,employee_name,"F",saturday,8,0)
                else:
                    upload_incideces(employee_number,employee_name,event_saturday,saturday,8,0)

def upload_incideces(employee_number,employee_name,incidence,date,hours,minutes):
    incidence_obj = Incidences()
    incidence_obj.number = employee_number
    incidence_obj.name = employee_name
    incidence_obj.incidence = incidence
    incidence_obj.date = date
    incidence_obj.hours = hours
    incidence_obj.minutes=minutes 
    incidence_obj.save()

def clean_decimals(number):
    if "." in number:
        number = number[:-2]
    return number

def upload_employees(ws,wb):
    '''
    This function uploads employees using an excel file
    with a specific format described below in the code
    '''
    #Personal variables
    rows = []
    offset = 1
    cont =0
    for i, row in enumerate(range(ws.nrows)):
            if i < offset:  # (Optionally) skip headers
                continue
            ne = str(ws.cell(i, 0).value)
            ne = ne[:-2]
            try:
                cont=cont+1
                emp = Employees.objects.get(master_id=ne)
                if emp:
                    load_employee(i,ws,wb,emp)
            
            except Exception as e:
                print(""+str(e))
                
                employee = Employees()
                load_employee(i,ws,wb,employee)

def load_employee(i,ws,wb,employee):
    id = str(ws.cell(i, 0).value)
    
    employee.master_id = id[:-2]
    old = str(ws.cell(i, 1).value)
    employee.old_number = old[:-2]
    numberx = str(ws.cell(i, 2).value)
    numberx = numberx[:-2]
    employee.number = numberx
    employee.name = ws.cell(i, 3).value
    employee.group = ws.cell(i, 4).value
    employee.team = ws.cell(i, 5).value
    employee.personal = ws.cell(i, 6).value
    employee.department_rp = ws.cell(i, 7).value
    employee.area_rp = ws.cell(i, 8).value
    employee.position_rp = ws.cell(i, 9).value
    employee.payroll = ws.cell(i, 10).value

    a1 = ws.cell_value(i, 11)
    a1_as_datetime = datetime.datetime(*xlrd.xldate_as_tuple(a1, wb.datemode))
    employee.admission_date = a1_as_datetime
    employee.antiquity = ws.cell(i, 12).value
    employee.status = ws.cell(i, 13).value
    a1 = ws.cell(i, 14).value
    if a1 == "NO ALTA":
        employee.leaving_date =  "NO ALTA"
    elif a1 != "":
        a1_as_datetime = datetime.datetime(*xlrd.xldate_as_tuple(a1, wb.datemode))
        employee.leaving_date =  a1_as_datetime
    else:
        employee.reason_of_leaving = ""
    employee.reason_of_leaving = ws.cell(i, 15).value
    employee.curp = ws.cell(i, 16).value
    employee.rfc = ws.cell(i, 17).value
    employee.nss = ws.cell(i, 18).value

    if ws.cell(i, 19).value != "":
        a1 = ws.cell(i, 19).value
        a1_as_datetime = datetime.datetime(*xlrd.xldate_as_tuple(a1, wb.datemode))
        employee.birth_date = a1_as_datetime
    else:
        employee.birth_date = ""
    
    employee.age = ws.cell(i, 20).value
    employee.gender = ws.cell(i, 21).value
    employee.department = ws.cell(i, 22).value
    employee.position = ws.cell(i, 23).value
    employee.jacket = ws.cell(i, 24).value
    employee.application = ws.cell(i, 25).value
    employee.performance = ws.cell(i, 26).value
    employee.nationality = ws.cell(i, 27).value
    employee.place_of_birth = ws.cell(i, 28).value
    employee.municipality = ws.cell(i, 29).value
    employee.location = ws.cell(i, 30).value
    employee.division = ws.cell(i, 31).value
    employee.suburb = ws.cell(i, 32).value
    employee.street = ws.cell(i, 33).value
    employee.house_number = ws.cell(i, 34).value
    employee.studies = ws.cell(i, 35).value
    employee.email = ws.cell(i, 36).value
    employee.phone = ws.cell(i, 37).value
    try:
        list = ws.cell(i, 38).value
        if list !="":
                if list.find('/') != -1:
                    value = list.split('/')
                    employee.blood_type = value[0]
                    employee.allergies = value[1]
                else:
                    if list.find('+') or list.find('-'):
                        employee.blood_type = list
                        employee.allergies = ""
                    else:
                        employee.blood_type = ""
                        employee.allergies = list
        else:
            employee.blood_type = ""
            employee.allergies = ""
    except Exception as e:
        print("Error"+str(e))

    employee.marital_status = ws.cell(i,39).value
    employee.route = ws.cell(i, 40).value
    employee.bus_stop = ws.cell(i, 41).value
    employee.emergency_phone_1 = ws.cell(i, 42).value
    employee.emergency_phone_2 = ws.cell(i, 43).value
    employee.emergency_phone_3 = ws.cell(i, 44).value
    employee.schedule = ws.cell(i, 45).value
    employee.number_of_children = ws.cell(i, 46).value
    employee.age_of_kid_1 = ws.cell(i, 47).value
    employee.age_of_kid_2 = ws.cell(i, 48).value
    employee.age_of_kid_3 = ws.cell(i, 49).value
    employee.age_of_kid_4= ws.cell(i, 50).value
    employee.age_of_kid_5 = ws.cell(i, 51).value
    employee.id_noi = ws.cell(i, 52).value
    employee.boss_support = ws.cell(i, 53).value
    employee.relation = ws.cell(i, 54).value
    employee.save()

def insert_on_excel(wb,ws,employee,row):
    
    c=0
    c=c+1
    ws.cell(row, c).value = employee.master_id
    ws.cell(1, 1).value = "MASTER ID"
    c=c+1
    ws.cell(row, c).value = employee.old_number
    ws.cell(1, 2).value = "Numero Antiguo"
    c=c+1
    ws.cell(row, c).value = employee.number
    ws.cell(1, 3).value = "Numero"
    c=c+1
    ws.cell(row, c).value = employee.name
    ws.cell(1, 4).value = "Nombre"
    c=c+1
    ws.cell(row, c).value = employee.group
    ws.cell(1, 5).value = "Grupo"
    c=c+1
    ws.cell(row, c).value = employee.team
    ws.cell(1, 6).value = "Equipo"
    c=c+1
    ws.cell(row, c).value = employee.personal
    ws.cell(1, 7).value = "Personal"
    c=c+1
    ws.cell(row, c).value = employee.department_rp
    ws.cell(1, 8).value = "Depto. RP"
    c=c+1
    ws.cell(row, c).value = employee.area_rp
    ws.cell(1, 9).value = "Area RP"
    c=c+1
    ws.cell(row, c).value = employee.position_rp
    ws.cell(1, 10).value = "Puesto RP"
    c=c+1
    ws.cell(row, c).value = employee.payroll
    ws.cell(1, 11).value = "NOMINA"
    c=c+1
    ws.cell(row, c).value = employee.admission_date
    todays_date = date.today()
    start_date = datetime. datetime(int(employee.admission_date[0:4]),int(employee.admission_date[5:7]),int(employee.admission_date[8:10]))
    end_date = datetime. datetime(todays_date.year, todays_date.month, todays_date.day)
    num_months = (end_date. year - start_date. year) * 12 + (end_date. month - start_date. month)
    ws.cell(1, 12).value = "ALTA"
    c=c+1
    ws.cell(row, c).value = employee.antiquity
    ws.cell(1, 13).value = "ANTIGUEDAD"
    c=c+1
    ws.cell(row, c).value = employee.status
    ws.cell(1, 14).value = "Estatus"
    c=c+1
    ws.cell(row, c).value = employee.leaving_date
    ws.cell(1, 15).value = "Fecha de baja"
    c=c+1
    ws.cell(row, c).value = employee.reason_of_leaving
    ws.cell(1, 16).value = "Causa de baja"
    c=c+1
    ws.cell(row, c).value = employee.curp
    ws.cell(1, 17).value = "CURP"
    c=c+1
    ws.cell(row, c).value = employee.rfc
    ws.cell(1, 18).value = "RFC"
    c=c+1
    ws.cell(row, c).value = employee.nss
    ws.cell(1, 19).value = "NSS"
    c=c+1
    ws.cell(row, c).value = employee.birth_date
    ws.cell(1, 20).value = "Fecha de Nacimiento"
    c=c+1
    ws.cell(row, c).value = employee.age
    ws.cell(1, 21).value = "Edad"
    c=c+1
    ws.cell(row, c).value = employee.gender
    ws.cell(1, 22).value = "Sexo"
    c=c+1
    ws.cell(row, c).value = employee.department
    ws.cell(1, 23).value = "Departamento"
    c=c+1
    ws.cell(row, c).value = employee.position
    ws.cell(1, 24).value = "Puesto"
    c=c+1
    ws.cell(row, c).value = employee.jacket
    ws.cell(1, 25).value = "Chaleco"
    c=c+1
    ws.cell(row, c).value = employee.aplication
    ws.cell(1, 26).value = "Aplicacion"
    c=c+1
    ws.cell(row, c).value = employee.performance
    ws.cell(1, 27).value = "Rendimiento"
    c=c+1
    ws.cell(row, c).value = employee.nationality
    ws.cell(1, 28).value = "Nacionalidad"
    c=c+1
    ws.cell(row, c).value = employee.place_of_birth
    ws.cell(1, 29).value = "Lugar de nacimiento"
    c=c+1
    ws.cell(row, c).value = employee.municipality
    ws.cell(1, 30).value = "Municipio"
    c=c+1
    ws.cell(row, c).value = employee.location
    ws.cell(1, 31).value = "Localidad"
    c=c+1
    ws.cell(row, c).value = employee.division
    ws.cell(1, 32).value = "Fraccionamiento"
    c=c+1
    ws.cell(row, c).value = employee.suburb
    ws.cell(1, 33).value = "Colonia"
    c=c+1
    ws.cell(row, c).value = employee.street
    ws.cell(1, 34).value = "Calle"
    c=c+1
    ws.cell(row, c).value = employee.house_number
    ws.cell(1, 35).value = "Numero Casa"
    c=c+1
    ws.cell(row, c).value = employee.studies
    ws.cell(1, 36).value = "Nivel de estudios"
    c=c+1
    ws.cell(row, c).value = employee.email
    ws.cell(1, 37).value = "E-mail"
    c=c+1
    ws.cell(row, c).value = employee.phone
    ws.cell(1, 38).value = "Télefono"
    c=c+1
    ws.cell(row, c).value = employee.blood_type+"/"+employee.allergies
    ws.cell(1, 39).value = "Tipo de sangre"
    c=c+1
    ws.cell(row, c).value = employee.marital_status
    ws.cell(1, 40).value = "Estado Civil"
    c=c+1
    ws.cell(row, c).value = employee.route
    ws.cell(1, 41).value = "Ruta"
    c=c+1
    ws.cell(row, c).value = employee.bus_stop
    ws.cell(1, 42).value = "Parada"
    c=c+1
    ws.cell(row, c).value = employee.emergency_phone_1
    ws.cell(1, 43).value = "No. Emergencia 1"
    c=c+1
    ws.cell(row, c).value = employee.emergency_phone_2
    ws.cell(1, 44).value = "No. Emergencia 2"
    c=c+1
    ws.cell(row, c).value = employee.emergency_phone_3
    ws.cell(1, 45).value = "No. Emergencia 3"
    c=c+1
    ws.cell(row, c).value = employee.schedule
    ws.cell(1, 46).value = "Horario"
    c=c+1
    ws.cell(row, c).value = employee.number_of_children
    ws.cell(1, 47).value = "Cantidad Hijos"
    c=c+1
    ws.cell(row, c).value = employee.age_of_kid_1
    ws.cell(1, 48).value = "Edad de hijo 1"
    c=c+1
    ws.cell(row, c).value = employee.age_of_kid_2
    ws.cell(1, 49).value = "Edad de hijo 2"
    c=c+1
    ws.cell(row, c).value = employee.age_of_kid_3
    ws.cell(1, 50).value = "Edad de hijo 3"
    c=c+1
    ws.cell(row, c).value = employee.age_of_kid_4
    ws.cell(1, 51).value = "Edad de hijo 4"
    c=c+1
    ws.cell(row, c).value = employee.age_of_kid_5
    ws.cell(1, 52).value = "Edad de hijo 5"
    c=c+1
    ws.cell(row, c).value = employee.id_noi
    ws.cell(1, 53).value = "Id Noi"
    c=c+1
    ws.cell(row, c).value = employee.boss_support
    ws.cell(1, 54).value = "Apoyo Jefe de Linea"
    c=c+1
    ws.cell(row, c).value = employee.relation
    ws.cell(1, 55).value = "Relacion"
    c=c+1

def create_payroll_excels():
    wb_1 = Workbook()
    sheet = wb_1.active
    sheet.cell(row=1, column=1).value = "Clave trabajador"
    sheet.cell(row=1, column=2).value = "Tipo Falta"
    sheet.cell(row=1, column=3).value = "% falta"
    sheet.cell(row=1, column=4).value = "Certificado IMSS"
    sheet.cell(row=1, column=5).value = "% pagado por IMSS"
    sheet.cell(row=1, column=6).value = "Fecha inicio"
    sheet.cell(row=1, column=7).value = "Fecha fin"
    sheet.cell(row=1, column=8).value = "Clave tipo de incidencia"
    sheet.cell(row=1, column=9).value = "Observaciones"
    wb_2 = Workbook()
    sheet = wb_2.active
    sheet.cell(row=1, column=1).value = "Clave trabajador"
    sheet.cell(row=1, column=2).value = "Per/Ded"
    sheet.cell(row=1, column=3).value = "Núm Per/Ded"
    sheet.cell(row=1, column=4).value = "Núm crédito"
    sheet.cell(row=1, column=5).value = "Aplica Destajo"
    sheet.cell(row=1, column=6).value = "Aplicación"
    sheet.cell(row=1, column=7).value = "Fecha inicio"
    sheet.cell(row=1, column=8).value = "Fecha fin"
    sheet.cell(row=1, column=9).value = "Monto o fórmula"
    sheet.cell(row=1, column=10).value = "Valor del descuento"
    sheet.cell(row=1, column=11).value = "Monto límite"
    sheet.cell(row=1, column=12).value = "Monto Acumulado"
    sheet.cell(row=1, column=13).value = "Criterio INFONAVIT"
    return wb_1,wb_2




def process_excel_date(cell_value,wb):
    day = datetime.datetime(*xlrd.xldate_as_tuple(cell_value, wb.datemode))
    return day


def last_employee(date_d):
    emps = Attendance.objects.filter(date=date_d).order_by('employee','date','time')
    num=0
    for emp in emps.iterator():
        pass
        num=emp.number
    return num
        


def create_day_attendance(date_d):
    pass
    attendance = Attendance.objects.filter(date=date_d).order_by('employee','date','time')
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "Fecha"
    ws['B1'] = str(date_d)
    ws['A2'] = "Area"
    ws['B2'] = "No.Empleado"
    ws['C2'] = "Nombre"
    ws['D2'] = "Falta Checadas"
    ws['E2'] = "Autorización por falta de checadas"
    ws['F2'] = "Firma"

    number = ""
    oldnumber = ""
    event_final=""
    ws_row = 3
    ws_column = 1
    eval0 = 0
    eval1 = 0
    eval2 = 0
    eval3 = 0
    eval4 = 0
    eval5 = 0
    eval6 = ""
    name1 = ""
    oldnumber = ""
    oldname = ""
    olddepartment = ""
    texto_checadas = ""
    texto_checadas2 = ""
    department1 = ""
    
    last = []

    numero_final = last_employee(date_d)
    for att in attendance.iterator():
        date = att.date
        time = att.time
        number1 = att.number
        event = att.event
        device = att.device
        employee = att.employee

        if eval6 == "" and number1 != oldnumber or number1 == numero_final:
            pass
            if numero_final == number1:
                pass
                last.append(event)
                

            oldnumber = number1
            oldname = employee
            olddepartment = department1
            eval6 = "1"

        elif number1 != oldnumber :
            emps2 = Employees.objects.filter(number=number1)
 
            for emp2 in emps2.iterator():
                department1 = emp2.department_rp
                
            if eval0 == 0:
                pass
                texto_checadas2=texto_checadas2+"Entrada - "
            if eval1 == 0:
                texto_checadas = texto_checadas+"Salida a descanso - "
            if eval2 == 0:
                texto_checadas = texto_checadas+"Regreso descanso - "
            if eval3 == 0:
                pass
                texto_checadas2 = texto_checadas2+"Salida - "
            if eval4 == 1 or eval5 == 1:
                if eval4 == 0:
                    texto_checadas = texto_checadas+"Entrada T.E. - "
                if eval5 == 0:
                        texto_checadas = texto_checadas+"Salida T.E. - "
            ws.cell(ws_row, ws_column).value = olddepartment
            ws.cell(ws_row, ws_column+1).value = oldnumber
            ws.cell(ws_row, ws_column+2).value = oldname
            ws.cell(ws_row, ws_column+3).value = texto_checadas
            if(texto_checadas != ""):
                ws.cell(ws_row, ws_column+4).value = 1
            ws.cell(ws_row, ws_column+6).value = texto_checadas2
            oldnumber = number1
            olddepartment = department1
            oldname = employee
            ws_row = 1+ws_row

            # Reseteo de variables
            eval0 = 0
            eval1 = 0
            eval2 = 0
            eval3 = 0
            eval4 = 0
            eval5 = 0
            texto_checadas = ""
            texto_checadas2 = ""
        # Evaluacion para imprimir el texto
        if event == "Entrada":
            eval0 = 1
        if event == "Salida a descanso":
            eval1 = 1
        if event == "Regreso descanso":
            eval2 = 1
        if event == "Salida":
            eval3 = 1
        if event == "Entrada T.E.":
            eval4 = 1
        if event == "Salida T.E.":
            eval5 = 1

    if(texto_checadas != ""):
        ws.cell(ws_row, ws_column+4).value = 1
    if "Salida a descanso" not in last:
        pass
        event_final = event_final+"Salida descanso - "
    if "Regreso descanso" not in last:
        event_final = event_final+"Regreso a descanso - "
    if eval4 == 1 or eval5 == 1:
        if 'Entrada T.E.' not in last:
            event_final = event_final+"Entrada T.E. - "
        if 'Salida T.E.' not in last:
            pass
            event_final = texto_checadas+"Salida T.E. - "
    ws.cell(ws_row, ws_column).value = olddepartment
    ws.cell(ws_row, ws_column+1).value = oldnumber
    ws.cell(ws_row, ws_column+2).value = oldname
    ws.cell(ws_row, ws_column+3).value = event_final
    wb.save("FaltaChecadas.xlsx")
     