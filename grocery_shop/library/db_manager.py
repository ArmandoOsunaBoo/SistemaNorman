
#Django imports
from django.shortcuts import render,redirect

#Personal imports
from employees.models import Employees, Attendance
import xlrd
import datetime
from django.http.response import HttpResponse
from datetime import date
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from PyPDF3 import PdfFileMerger
from grocery_shop.models import Product,Category,Unit,Shop,Order
from users.models import User
from datetime import datetime as dt,date, timedelta
import xlwt



def update_cart():
    today = date.today()
    actual_date = today.strftime("%Y-%m-%d")
    actual_order = Order()
    actual_order.date = actual_date


def get_cart(user_id):
    first_date,seccond_date = dates_range()
    products = Order.objects.filter(date__range=(first_date,seccond_date),id_user=user_id)
    return products


def insert_order(request):
    pass
    product_name = request.POST['name']
    id_product = request.POST['product_id']
    product = Product.objects.get(id=id_product)
    unit=request.POST['unit']
    quantity = request.POST['quantity']
    user_name = request.POST['user_name']
    product_unit = request.POST['unit']
    product_price = request.POST['price']
    first_day,seccond_day = dates_range()
    user = User.objects.get(id=int(request.POST['user_id'])) 
    user_id = user
    today = date.today()
    actual_date = today.strftime("%Y-%m-%d")
    try:
        if Order.objects.filter(date__range=(first_day,seccond_day),id_user=user.id,product_name=product_name).exists():
            order = Order.objects.get(date__range=(first_day,seccond_day),id_user=user.id,product_name=product_name)
            quantity = float(quantity) + float(order.amount)
            save_order(order,actual_date,quantity,product_unit,product_name,user_name,product_price,user_id,id_product)
            message = "¡Compra Actualizada Correctamente!"
            return message,unit,product
        else:
            order = Order()
            save_order(order,actual_date,quantity,product_unit,product_name,user_name,product_price,user_id,id_product)    
            message = "¡Producto Guardado Correctamente!"
    except Exception as e:
        pass
        print("GOT AN ERROR")
        print(e)
        message = "¡Wops No se pudo guardar tu compra! :("
    finally:
        pass
        return message,unit,product
    




def upload_products_by_excel(excel_file):
    pass
    #utility
    offset = 1
    # you may put validations here to check extension or file size
    wb = xlrd.open_workbook(excel_file.temporary_file_path())
    # getting a particular sheet by name out of many sheets
    ws = wb.sheet_by_index(0)
    all_categories = Category.objects.all()
    all_shops = Shop.objects.all()
    all_units = Unit.objects.all()
    all_products = Product.objects.all()
    obj = Product()
    for i, row in enumerate(range(ws.nrows)):
            if i < offset:  
                continue
            r = []
            master = ws.cell(i, 0).value
            if not Category.objects.filter(name=ws.cell(i, 7).value).exists():
                pass
                obj = Category(name=ws.cell(i, 7).value,name_ch=ws.cell(i, 7).value)
                obj.save()
            if not Unit.objects.filter(name=ws.cell(i, 12).value).exists():
                pass
                obj = Unit(name=ws.cell(i, 12).value,name_ch =ws.cell(i,12).value)
                obj.save()
            if not Shop.objects.filter(name=ws.cell(i, 8).value).exists():
                pass
                obj = Shop(name=ws.cell(i, 8).value, name_ch =ws.cell(i, 8).value)
                obj.save()
                
            if Product.objects.filter(id_master=master).exists():
               pass
               #need to update the product
               obj = Product.objects.get(id_master=master)
               create_products(ws,i,obj)
            else:
                obj = Product()
                create_products(ws,i,obj)

def create_products(ws,i,obj):
    pass
    p=str(ws.cell(i,0).value)
    obj.id_master = p[:-2]
    obj.name = ws.cell(i, 2).value
    obj.name_ch = ws.cell(i, 3).value
    obj.description = ws.cell(i, 5).value
    obj.description_ch = ws.cell(i, 6).value
    p = str(ws.cell(i, 10).value)
    obj.price = p[:-2]
    print(p[:-2])
    obj.code = ws.cell(i, 9).value
    obj.image = ws.cell(i, 11).value+".jpg"
    objc = Category.objects.get(name=ws.cell(i, 7).value)
    obj.category_name_id = objc.id
    objc = Shop.objects.get(name=ws.cell(i, 8).value)
    obj.shop_name_id = objc.id
    objc = Unit.objects.get(name=ws.cell(i, 12).value)
    obj.unit_name_id = objc.id
    obj.save()



def dates_range():
    day =["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    actual_date  = dt.now()
    x=actual_date
    first_day=""
    seccond_day=""
    offset = (x.weekday() - 1) % 7
    if(day[x.weekday()]=="Tuesday"):
        pass
        print("--------------------------hoy es martes")
        seccond_day = x + timedelta(days=x.weekday(), weeks=0)
        first_day = x + timedelta(days=1-x.weekday(), weeks=-1)
    else:
        print("--------------------------hoy NO es martes")
        if(day[x.weekday()]=="Monday"):
            print("Hoy  es lunes")
            first_day = x + timedelta(days=1-x.weekday(), weeks=-1)
            seccond_day = x + timedelta(days=x.weekday(), weeks=0)
        else:
            print("Hoy NO es lunes")
            first_day = x - timedelta(days=offset)
            seccond_day = x 
    return first_day,seccond_day



def save_order(order,actual_date,quantity,product_unit,product_name,user_name,product_price,user_id,id_product):
    order.date = actual_date
    order.amount = quantity
    order.unit = product_unit
    order.product_name = product_name
    order.user_name = user_name
    order.unidad_cantidad = product_unit
    order.unit_price = product_price
    order.id_user = user_id
    product = Product.objects.get(id=id_product)
    order.product_ref = product
    order.save()





def generate_individual_reports(week):
        print("*****************************************")
        week2=str(week)
        week2=week2[-2:]
        week2=int(week2)
        start_thuesday = datetime.datetime.strptime(week + '-2', "%Y-W%W-%w")
        start_thuesday= str(start_thuesday)
        today = datetime.datetime(int(start_thuesday[0:4]), int(start_thuesday[5:7]), int(start_thuesday[8:10]))
        next_monday = today + datetime.timedelta( (7-today.weekday()) % 7 )
        orders = Order.objects.filter(date__range=(today, next_monday)).order_by('-id_user')
        wb = Workbook()
        ws = wb.active
        row=1
        column=1
        p_row=4
        p_column=1
        actual_total=0.0
        old_user=""
        for order in orders.iterator():
            actual_user=order.id_user

            if actual_user !=old_user and old_user!="":
                ws.cell(p_row,p_column).value = "TOTAL"
                ws.cell(p_row,p_column+1).value = "$ "+str(actual_total)
                actual_total=0
                p_row=4
                p_column=p_column+4
                row=1
                column=column+4
            old_user=actual_user
            ws.cell(row,column).value = "Nombre"
            ws.cell(row,column+1).value = order.id_user.username
            ws.cell(row+1,column).value = "PRODUCTO"
            ws.cell(row+1,column+1).value = "CANTIDAD"


            ws.cell(p_row,p_column).value = order.product_name
            ws.cell(p_row,p_column+1).value = order.amount
            ws.cell(p_row,p_column+2).value = order.unit
            actual_total = actual_total + (float(order.unit_price)*float(order.amount))

            p_row=p_row+1

        ws.cell(p_row,p_column).value = "TOTAL"
        ws.cell(p_row,p_column+1).value = "$ "+str(actual_total)

        wb.save("ListadoCompleto.xlsx")
        fh = open("ListadoCompleto.xlsx", 'rb')
        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'inline; filename=' + os.path.basename('ListadoCompleto.xlsx')
        return response
            


def generate_group_reports(week):
        week2=str(week)
        week2=week2[-2:]
        week2=int(week2)
        start_thuesday = datetime.datetime.strptime(week + '-2', "%Y-W%W-%w")
        start_thuesday= str(start_thuesday)
        today = datetime.datetime(int(start_thuesday[0:4]), int(start_thuesday[5:7]), int(start_thuesday[8:10]))
        next_monday = today + datetime.timedelta( (7-today.weekday()) % 7 )
        shops = Shop.objects.all()
        old_shop=""
        actual_shop=""
        actual_product=""
        old_product=""
        actual_amount=0.0
        old_amount=0.0
        wb = Workbook()
        row=3
        column=2
        actual_product_name=""
        old_product_name=""
        # grab the active worksheet
        ws = wb.active
        s_row=2
        s_column=2
        total_price=0
        for shop in shops.iterator():
            print("****************************")
            print(shop.name)
            ws.cell(s_row,s_column).value = "TIENDA"
            ws.cell(s_row,s_column+1).value = shop.name
            ws.cell(s_row+1,s_column).value = "PRODUCTO"
            ws.cell(s_row+1,s_column+1).value = "CANTIDAD"
            orders = Order.objects.filter(product_ref__shop_name=shop.id,date__range=(today, next_monday)).order_by('-product_ref__shop_name','-product_ref_id')
            print(orders.count())
            s_row=s_row+2
            for order in orders.iterator():
                actual_shop =  order.product_ref.shop_name.name
                actual_product = order.product_ref.id
                actual_product_name = order.product_ref.name
                actual_amount = float(order.amount)
                print("patapatapatapon")
                print(actual_product_name)
                print(actual_amount)
                total_price = total_price + (float(order.unit_price)*float(actual_amount))
                
                
                if old_product_name!=actual_product_name:
                    ws.cell(s_row,s_column).value = actual_product_name
                    ws.cell(s_row,s_column+1).value = actual_amount
                    ws.cell(s_row,s_column+2).value = order.unit
                    old_amount= actual_amount
                else:
                    s_row=s_row-1
                    ws.cell(s_row,s_column+1).value = actual_amount +old_amount
                    old_amount= actual_amount+old_amount

                old_product_name=actual_product_name
                s_row=s_row+1
            
            ws.cell(s_row,s_column).value = "TOTAL "
            ws.cell(s_row,s_column+1).value ="$"+str(total_price)
            ws.cell(s_row,s_column).value = "MXN "
            s_row = 2
            s_column= s_column+4
            total_price=0.0
        wb.save("ListadoCompleto.xlsx")
        fh = open("ListadoCompleto.xlsx", 'rb')
        response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
        response['Content-Disposition'] = 'inline; filename=' + os.path.basename('ListadoCompleto.xlsx')
        return response
                
""" 
print("/////////////////ACTUALES////////////////////")
print(actual_shop)
print(actual_product_name)
print(order.amount)
print("--------------VIEJOS------------------")
print(old_shop)
print(old_product_name)
print(old_amount) """


""" 

        if cont==1:
            old_shop = actual_shop
            old_product = actual_product
            old_product_name = actual_product_name
            old_amount = actual_amount 
            ws.cell(s_row,s_column).value = old_product_name
            ws.cell(s_row,s_column+1).value = old_amount
        else:
            pass
            if old_product == actual_product:
                print("IGUALES")
                old_amount = actual_amount + old_amount
                ws.cell(s_row,s_column+1).value = old_amount
                print("^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^^")
            else:
                print("DIFERENTES")
                ws.cell(s_row,s_column).value = old_product_name
                ws.cell(s_row,s_column+1).value = old_amount
                
                old_shop= actual_shop
                old_product = actual_product 
                old_product_name =  actual_product_name 
                old_amount= actual_amount
                s_row=s_row+1
    print("Saltamos a otra tienda")
    s_row = 2
    s_column= s_column+3 """

        
       






def generate_reports2(week):
        print("*****************************************")
        print(week)
        week2=str(week)
        week2=week2[-2:]
        week2=int(week2)
        print(week2)
        start_thuesday = datetime.datetime.strptime(week + '-2', "%Y-W%W-%w")
        start_thuesday= str(start_thuesday)
        today = datetime.datetime(int(start_thuesday[0:4]), int(start_thuesday[5:7]), int(start_thuesday[8:10]))
        print(today)
        next_monday = today + datetime.timedelta( (7-today.weekday()) % 7 )
        print(next_monday)
        orders = Order.objects.filter(date__range=(today, next_monday)).order_by('-product_ref__shop_name')
        workbook = xlwt.Workbook()
        print("-----------------------------------------")
        for order in orders.iterator():
            print(order.product_ref.shop_name.name)



def insert_order_ajax(request,order_id,user,product):
    pass
    product_name = product.name
    print(product_name)
    id_product = product.id
    unit=product.unit_name.name 
    user_name = user.username
    product_unit = product.unit_name.name
    product_price = product.price
    first_day,seccond_day = dates_range()
    user_id = user
    today = date.today()
    actual_date = today.strftime("%Y-%m-%d")
    try:
        if Order.objects.filter(date__range=(first_day,seccond_day),id_user=user.id,product_name=product_name).exists():
            order = Order.objects.get(date__range=(first_day,seccond_day),id_user=user.id,product_name=product_name)
            quantity = float(order.amount) + 1
            save_order(order,actual_date,quantity,product_unit,product_name,user_name,product_price,user_id,id_product)
            message = "¡Compra Actualizada Correctamente!"
            return message,unit,product
        else:
            order = Order()
            save_order(order,actual_date,1,product_unit,product_name,user_name,product_price,user_id,id_product)    
            message = "¡Producto Guardado Correctamente!"
    except Exception as e:
        pass
        message = "¡Wops No se pudo guardar tu compra! :("
    finally:
        pass
        return message,unit,product